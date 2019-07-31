#encoding:utf-8
import os, re, sys
import numpy as np
from datetime import datetime, timedelta, time
from openpyxl import Workbook,load_workbook

from icalendar import Calendar, Event, LocalTimezone, vDatetime
from datetime import datetime, timedelta, time
from sys import platform

time_spans = [ time(9,0,0), time(10,45,0), time(13,0,0), time(14,45,0), time(16,30,0), time(18,15,0), time(20,00,0) ]
lab_timespans = [ time(9,0,0), time(13,0,0), time(16,30,0) ] # , time(18,15,0)

day_names = [ u'Понедельник', u'Вторник', u'Среда', u'Четверг', u'Пятница', u'Суббота' ]
this_year = 2019
fio_re = re.compile( u'[А-Яа-я]* [А-Я]\.[А-Я]\.' )
hourspan_re = re.compile( u"[0-9]{2}:[0-9]{2}-[0-9]{2}:[0-9]{2}" )

day_rows = [5,7,9,11,13,15]
hour_spans = [2,3,4,5,6,7,8]

# Сократить название кабинета
s_nb = u'⁰¹²³⁴⁵⁶⁷⁸⁹'
def s_room( s ):
    s_out = s;
    for i,j in enumerate(s_nb): s_out = s_out.replace( '(%s)'%str(i), j )
    s_out = s_out.replace( u'(ГУК А)', u'ГУКᴬ' )
    s_out = s_out.replace( u'(ГУК Б)', u'ГУК⁶' )
    s_out = s_out.replace( u'(ГУК В)', u'ГУКᴮ' )
    return s_out

def shorten( subj_words ):
  short_name = ""
  if len(subj_words) == 1:
  	if len(subj_words[0]) > 7: return subj_words[0][:7]+'.'
  	else: return subj_words[0]
  for s in subj_words:
    t = s.replace('(','').replace(')','')
    if len(t) == 1: short_name = short_name + t
    elif t.isupper(): short_name = short_name + t
    else: short_name = short_name + t[0].upper()
  if len( short_name ) <= 2:
    short_name = ''
    for s in subj_words:
      t = s.replace('(','').replace(')','')
      if len(t) == 1: short_name = short_name + t
      elif t.isupper(): short_name = short_name + t
      else: short_name = short_name + t[0].upper() + t[1:3]
  return short_name

def shorten_group_name( g ):
  # Простой вариант
  p = u''
  if len(g)>1: p = u'…'
  s = u','.join( g )
  compiled = s.replace('(',' ').replace(',',' ').split(' ')[0]
  #if '-' in compiled: compiled = compiled.split(u'-')[1]
  return compiled + p



#files = [f for f in os.listdir( os.path.join(sys.argv[1],'xlsx') ) if os.path.isfile(os.path.join(sys.argv[1],'xlsx')) and f.split('.')[-1]==u'xlsx']

#for f in os.listdir( os.path.join(sys.argv[1],'xlsx')): print os.path.join(sys.argv[1],'xlsx',f)
files = [f for f in os.listdir( os.path.join(sys.argv[1],'xlsx') ) if os.path.isfile(os.path.join(sys.argv[1],'xlsx',f)) and f.split('.')[-1]==u'xlsx']
#for f in files: print f
#exit()

# Считать события сначала?
sem_start = datetime( year=2019, month=9, day=2 )
sem_finish = datetime( year=2019, month=12, day=31)

print u'Начало семестра:', sem_start.strftime( u"%d.%m.%Y" )
print u'Конец семестра:', sem_finish.strftime( u"%d.%m.%Y" )
print ''
firstweek_nb = sem_start.isocalendar()[1]



total_cal = Calendar()
total_cal.add('version', '2.0')
total_cal.add('X-WR-CALNAME', "Общий календарь кафедры" )



for f in sorted(files):

  if platform == "linux" or platform == "linux2":
    print '%', f
  elif platform == "win32":
    print f.decode('cp1251')

  wb = load_workbook(filename = os.path.join(sys.argv[1],'xlsx',f))

  #ws = wb.active # grab the active worksheet
  ws = wb['TDSheet'] # grab the active worksheet
  #merged = [ s[:s.find(':')] for s in ws.merged_cell_ranges ]

  #print ' '
  #for w in ws.merged_cells.ranges:
  #  print str(w)[ :str(w).find(':') ]
  #  #print dir(w)
  #  print ws.cell(row=w.min_row, column=w.min_col)
  #print ' '

  #merged = [ s[:s.find(':')] for s in ws.merged_cells.ranges ]
  merged = [ str(w)[ :str(w).find(':')] for w in ws.merged_cells.ranges ]

  if u'преподавателя' in ws['C2'].value: type_of_record = u'преподаватель'; name = ' '.join(ws['C2'].value.split(' ')[-2:-1])
  elif u'группы' in ws['C2'].value: type_of_record = u'группа'; name = (ws['C2'].value[  ws['C2'].value.find(u'группы')+len(u"группы"):  ]).strip()
  elif u'аудитории' in ws['C2'].value: type_of_record = u'аудитория';  name = (ws['C2'].value[  ws['C2'].value.find(u'аудитории')+len(u"аудитории"):  ]).strip()

  # Создаём новый календарь
  cal = Calendar()
  cal.add('version', '2.0')
  #cal.add('prodid', '-//test file//example.com//')
  cal.add('X-WR-CALNAME', name )
  #lt = LocalTimezone() # we append the local timezone to each time so that icalendar will convert to UTC

  #print type_of_record.encode('utf-8')
  #print name.encode('utf8')

  for ir,r in enumerate(day_rows):
    for ic,c in enumerate(hour_spans):
      for u in [0,1]:
        if ws.cell(column=c,row=r+u).value != None:

          for rec in ws.cell(column=c,row=r+u).value.split('---\n'):

            #if u"Капырин" in name: print time_spans[ic], rec
            #if u"Мишин" in name: print time_spans[ic], rec
            #if u"Кузнецов" in name: print time_spans[ic], rec

            if ws.cell(column=c,row=r+u).coordinate in merged: updn = u''
            elif u == 0: updn = u'в.н.'
            elif u == 1: updn = u'н.н.'

            #print rec.replace('\n','//')

            if type_of_record == u'преподаватель':
              rec = rec.replace('\n','')
              rec = rec[:rec.find(u',')] + u'|' + rec[rec.find(u',')+1:]
              rec = rec.replace(u', ЛР,', u'|ЛБ|').replace(u', ЛК,', u'|ЛК|').replace(u', ПЗ,', u'|ПЗ|').replace(u', КСР,', u'|КСР|')
              #rec = rec.replace(u', ЛР,', u'|лаб.|').replace(u', ЛК,', u'|лекция|').replace(u', ПЗ,', u'|практика|').replace(u', КСР,', u'|КСР|')
              prep_name = name
              data = rec.split('|')
              room = s_room( data[0].strip().replace(u'ауд.',u'').replace(u'Ауд.',u'') )
              subj = data[1].strip()
              htype = data[2].strip()
              tmp = data[3].split('(')
              subj_words = subj.replace(',',' ').replace('-',' ').replace('  ',' ').split(' ')
              short_name = shorten( subj_words )
              groups = [ t.strip(' ') for t in tmp[0].split(',') ]
              span = tmp[1].strip(' ').strip(')')

            elif type_of_record == u'группа':
              rec = rec.replace(u', ЛР', u', ЛБ')
              #print rec.replace('\n','//')
              groups = [name]
              #rec = rec.replace(u', ЛР,', u'|ЛР|').replace(u', ЛК,', u'|ЛК|').replace(u', ПЗ,', u'|ПЗ|').replace(u', КСР,', u'|КСР|')
              data = rec.split('\n')
              if len(data)>2:
                  prep_name = data[0].strip()
                  prep_name = prep_name.split(' ')[0]
                  room = re.sub( hourspan_re, '', data[1]).strip()
                  room = s_room( room.strip().replace(u'ауд.',u'').replace(u'Ауд.',u'') )
                  subj_words = data[2][ :data[2].rfind(',') ].replace(',',' ').replace('-',' ').replace('  ',' ').split(' ')
                  short_name = shorten( subj_words )
                  htype = data[2][ :data[2].rfind('(') ].split(',')[-1].strip()
                  #tmp = data[2][ data[2].rfind(',')+1: ]
                  #htype = tmp[ :tmp.find('(') ].strip()
                  #htype = htype.replace(u'ЛР', u'лаб.').replace(u'ЛК', u'лекция').replace(u'ПЗ', u'практика').replace(u'КСР', u'КСР')
                  span = data[2][ data[2].rfind('('): ].strip().strip('()')
              else:
                  prep_name = ''
                  room = re.sub( hourspan_re, '', data[0]).strip()
                  room = s_room( room.strip().replace(u'ауд.',u'').replace(u'Ауд.',u'') )
                  subj_words = data[1][ :data[1].rfind(',') ].replace(',',' ').replace('-',' ').replace('  ',' ').split(' ')
                  short_name = shorten( subj_words )
                  htype = data[1][ :data[1].rfind('(') ].split(',')[-1].strip()
                  span = data[1][ data[1].rfind('('): ].strip().strip('()')

            elif type_of_record == u'аудитория':
              rec = rec.replace(u', ЛР', u', ЛБ')
              #rec = rec.replace(u', ЛР,', u'|ЛБ|').replace(u', ЛК,', u'|ЛК|').replace(u', ПЗ,', u'|ПЗ|').replace(u', КСР,', u'|КСР|')
              #print rec.replace('\n','//')
              room = name
              data = rec.split('\n')
              prep_name = re.sub( hourspan_re, '', data[0]).strip()
              prep_name = prep_name.split(' ')[0]
              subj_words = data[1][ :data[1].rfind(',') ].replace(',',' ').replace('-',' ').replace('  ',' ').split(' ')
              short_name = shorten( subj_words )
              htype = data[1][ data[1].rfind(',')+1: ].strip()
              #htype = htype.replace(u'ЛР', u'лаб.').replace(u'ЛК', u'лекция').replace(u'ПЗ', u'практика').replace(u'КСР', u'КСР')
              #htype = tmp[ :tmp.find('(') ].strip()
              #tmp = data[1][ data[1].rfind(',')+1: ]
              tmp = data[2][ :data[2].rfind('(') ].strip()
              groups = [ t.strip(' ') for t in tmp.split(',') ]
              span = data[2][ data[2].rfind('('): ].strip().strip('()')


            #if u"Кузнецов" in name:
            #  print '  > room', room
            #  print '  > short_name', short_name
            #  print '  > span', span
            #  print '  > groups', groups
            #  print '  > htype', htype

            #print 'prep_name', prep_name
            #print 'span', span
            #print 'room', room
            #print 'short_name', short_name
            #print 'htype', htype

            # Не будем создавать событие если это вторая запись про лабораторную
            #if htype == u'ЛБ' and time_spans[ic] not in lab_timespans: continue;
            #if htype==u'ЛБ': hlen = timedelta( minutes = 190 )
            #else: hlen = timedelta( minutes = 90 )

            if '-' in span: course_start = datetime.strptime( span.split('-')[0].strip(' ')+'.%d'%this_year, '%d.%m.%Y' )
            elif ',' in span: course_start = datetime.strptime( span.split(',')[0].strip(' ')+'.%d'%this_year, '%d.%m.%Y' )
            else: course_start = datetime.strptime( span+'.%d'%this_year, '%d.%m.%Y' )

            #if room == "3-304":
            #  print htype, time_spans[ic]

            if '-' in span: course_finish = datetime.strptime( span.split('-')[1].replace(')','').strip(' ')+'.%d'%this_year, '%d.%m.%Y' )
            elif ',' in span: course_finish = datetime.strptime( span.split(',')[1].replace(')','').strip(' ')+'.%d'%this_year, '%d.%m.%Y' )
            else: course_finish = course_start

            if course_finish < course_start: course_finish = datetime( year = course_finish.year+1, month=course_finish.month, day=course_finish.day )

            csdt = datetime( year=course_start.year, month=course_start.month, day=course_start.day, hour=time_spans[ic].hour, minute=time_spans[ic].minute )
            hlen = timedelta( minutes = 90 )
            cfdt = csdt + hlen

            #if u"Кузнецов" in name: print '  + ', csdt, '\n  + ', cfdt

            # Попробуем поискать предыдущее лабораторное занятие и объединить, если есть
            if htype == u'ЛБ': #  or short_name==u"ВоеПод"
              found_ev = None
              for ev in cal.walk():
                if ev.name == "VEVENT":
                  #if (ev['type'] == u'ЛБ' or ev['course']==u"ВоеПод") and \
                  if ev['type'] == u'ЛБ' and \
                      ev['course'] == short_name and \
                      ev['location'] == room and \
                      ev['prof'] == prep_name and \
                      ev['date_time_start'].day == course_start.day and \
                      ev['date_time_start'].month == course_start.month and \
                      ev['date_time_start'].hour == (csdt - timedelta( minutes = 105 )).hour:
                    found_ev = ev;
                    break;
              if found_ev:
                #print '>', found_ev['prof'], found_ev['course'], found_ev['location'], found_ev['type'], found_ev['date_time_start']
                #print '>', prep_name,        short_name,         room,                 htype,            csdt
              	found_ev['date_time_end'] = cfdt
                found_ev['dtend'] = vDatetime(cfdt).to_ical()
              	continue;

            if updn == '': days = 7
            else: days = 14
            span_btw_classes = timedelta( days = days )
            n_occurences = int(np.ceil((course_finish - course_start).days/float(days)))+1

            all_event_executions = [ course_start + k*span_btw_classes for k in range(0,n_occurences) if course_start + k*span_btw_classes <= min(course_finish,sem_finish) ]
            all_days_txt = [ d.strftime('%d:%m') for d in all_event_executions]
            all_week_nbs = [ d.isocalendar()[1] - firstweek_nb + 1 for d in all_event_executions ]
            for i in range(len(all_week_nbs)):
                if all_week_nbs[i]<0:
                    all_week_nbs[i]=all_week_nbs[i]+52
            all_week_nbs_str = [str(s) for s in all_week_nbs]

            #print u'; '.join( [ day_names[ir], time_spans[ic].strftime(u'%H:%M'), room, short_name, htype, u','.join(groups), u"недели " + ','.join(all_week_nbs)] ).encode('utf8')
            if platform == "linux" or platform == "linux2":
              print u'; '.join( [ prep_name, day_names[ir], time_spans[ic].strftime(u'%H:%M'), room, short_name, htype, u','.join(groups), u"недели " + ','.join(str(a) for a in all_week_nbs)] ).encode('utf8')
            elif platform == "win32":
              print_data = u'; '.join( [ prep_name, day_names[ir], time_spans[ic].strftime(u'%H:%M'), room, short_name, htype, u','.join(groups), u"недели " + ','.join(str(a) for a in all_week_nbs)] ).encode('utf8')
              print print_data.decode('utf-8').encode('cp1251','replace').decode('cp1251')

            ##########################################################
            # Здесь начинается неработющий код по работе с СОБЫТИЯМИ #
            event = Event()
            event_name = short_name + ' (' + htype + ')'

            if updn == '': repeat_interval = 1
            else: repeat_interval = 2
            event['rrule'] = "FREQ=WEEKLY;INTERVAL=%d;COUNT=%d" % (repeat_interval, len(all_week_nbs))

            event['updown'] = updn
            event['week_span_str'] = str(min(all_week_nbs)) + '-' + str(max(all_week_nbs))
            event['summary'] = short_name + '(%s:%s)' % (htype, shorten_group_name( groups ))
            event['course'] = short_name
            event['date_time_start'] = csdt
            event['dtstart'] = vDatetime(csdt).to_ical()
            event['date_time_end'] = cfdt
            event['dtend'] = vDatetime(cfdt).to_ical()
            #event['id'] = info['id']
            event['prof'] = prep_name
            event['location'] = room
            event['groups'] = shorten_group_name( groups )
            # *.cal is limited to 64 symbol columns, then it inserts a line break
            event['group'] = u' '.join( groups ) #.replace('\n','').replace('\r','')
            event['type'] = htype
            event['week_numbers'] = ','.join(all_week_nbs_str)
            #event['first'] = first_event
            cal.add_component( event )
            total_cal.add_component( event )
            ##########################################################


              # cal.add_component( event

              #rec = rec.replace(u', ЛР', u'|лаб.|').replace(u', ЛК', u'|лекция|').replace(u', ПЗ', u'|практика|').replace(u', КСР', u'|КСР|') # Для расписаний групп
              #if fio_re.search( rec ):
              #  if fio_re.search( rec ).start() == 0: print 'Yey!'; exit()
              #print rec

            if platform == "linux" or platform == "linux2":
              path = u'cal/' + unicode(f.strip('.xslx').decode('utf-8')) + u'.cal'
            elif platform == "win32":
              path = u'cal/' + unicode(f.strip('.xslx').decode('cp1251')) + u'.cal'  #ISO-8859-1

  ff = open( os.path.join( unicode(sys.argv[1]), path), 'wb')
  s = cal.to_ical().replace('\\;',';' ).replace('\\,',',' )
  ff.write( s )
  ff.close()

  print ''.encode('utf8')

ff = open( os.path.join(sys.argv[1], 'cal/all.cal'), 'wb' )
ff.write( total_cal.to_ical().replace('\\;',';' ).replace('\\,',',' ) )
ff.close()
